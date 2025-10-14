import openpyxl
from openai import OpenAI
from pathlib import Path
import time
import os

# === CONFIGURATION ===
API_KEY = "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"
MODEL_NAME = "gpt-4.1"

# EXCEL_DIR = "excels/long"
# PROMPT_TEMPLATE_FILE = "prompt_long.txt"
# OUTPUT_PROMPTS_FILE = "generated_prompts_long.txt"
# OUTPUT_LLM_BASE_DIR = "llm_outputs_long"

EXCEL_DIR = "excels/short"
PROMPT_TEMPLATE_FILE = "prompt_short.txt"
OUTPUT_PROMPTS_FILE = "generated_prompts_short.txt"
OUTPUT_LLM_BASE_DIR = "llm_outputs_short"

# === INITIALIZE CLIENT ===
client = OpenAI(api_key=API_KEY)

# === READ PROMPT TEMPLATE ===
with open(PROMPT_TEMPLATE_FILE, "r", encoding="utf-8") as f:
    prompt_template = f.read()

# === FIND ALL .xlsx FILES ===
excel_files = [
    # "angika_stories.xlsx",
    # "assamese_stories.xlsx",
    # "awadhi_stories.xlsx",
    # "bagheli_stories.xlsx",
    # "bhojpuri_stories.xlsx",
    # "braj_stories.xlsx",
    # "bundeli_stories.xlsx",
    # "dogri_stories.xlsx",
    # "hindi_stories.xlsx",
    # "konkani_stories.xlsx",
    # "magahi_stories.xlsx",
    # "maithili_stories.xlsx",
    # "nepali_stories.xlsx",
    # "pali_stories.xlsx",
    # "urdu_stories.xlsx",
    "telugu_stories.xlsx",
]

# === CHECK WHICH LANGUAGES ARE ALREADY PROCESSED ===
os.makedirs(OUTPUT_LLM_BASE_DIR, exist_ok=True)
completed_languages = {
    Path(f).stem
    for f in os.listdir(OUTPUT_LLM_BASE_DIR)
    if f.endswith(".txt")
}

print(f"\n Already processed languages: {sorted(completed_languages)}")

# === LOOP OVER EXCEL FILES ===
for workbook_file in excel_files:
    workbook_path = os.path.join(EXCEL_DIR, workbook_file)
    workbook_path = Path(workbook_path)
    language = workbook_path.stem.split("_")[0]

    # === SKIP IF ALREADY PROCESSED ===
    if language in completed_languages:
        print(f" Skipping {language} — already processed.")
        continue

    print(f"\n Processing workbook: {workbook_path} (Language: {language})")
    wb = openpyxl.load_workbook(workbook_path)

    # Create per-language output file
    output_llm_file = os.path.join(OUTPUT_LLM_BASE_DIR, f"{language}.txt")

    workbook_prompts = []
    workbook_outputs = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n Processing sheet: {sheet_name}")

        data = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            field, value = row
            if field and value:
                data[field.strip()] = str(value).strip()

        story = data.get("Story", "")
        question = data.get("Question", "")
        reference = data.get("Reference Answer", "")
        answers_keys = sorted(
            [k for k in data.keys() if k.startswith("Answer_")],
            key=lambda x: int(x.split("_")[1])
        )
        answer_texts = "\n".join([f"{k}: {data[k]}" for k in answers_keys])

        # === Construct Prompt ===
        prompt = prompt_template.format(
            Language=language,
            Story=story,
            Question=question,
            Reference=reference,
            Answer_Texts=answer_texts
        )
        workbook_prompts.append((workbook_path.name, sheet_name, prompt))

        # === Call LLM API with Retry on Rate Limit ===
        # while True:
        #     try:
        #         response = client.chat.completions.create(
        #             model=MODEL_NAME,
        #             messages=[{"role": "user", "content": prompt}]
        #         )
        #         llm_output = response.choices[0].message.content
        #         workbook_outputs.append((workbook_path.name, sheet_name, llm_output))
        #         print(f"✅ LLM output received for {sheet_name}")
        #         break
        #     except Exception as e:
        #         error_str = str(e)
        #         if "rate limit" in error_str.lower():
        #             print("⚠️ Rate limit hit — retrying in 60 seconds...")
        #             time.sleep(60)
        #             continue
        #         else:
        #             print(f" Error generating output for {sheet_name}: {e}")
        #             workbook_outputs.append((workbook_path.name, sheet_name, f"Error: {e}"))
        #             break

        # time.sleep(12)  # small cooldown

    # === WRITE RESULTS TO FILE AFTER EACH WORKBOOK ===
    with open(OUTPUT_PROMPTS_FILE, "a", encoding="utf-8") as f:
        for workbook_name, sheet_name, prompt in workbook_prompts:
            f.write(f"=== Workbook: {workbook_name} | Sheet: {sheet_name} ===\n{prompt}\n{'='*80}\n\n")

    with open(output_llm_file, "a", encoding="utf-8") as f:
        for workbook_name, sheet_name, output in workbook_outputs:
            f.write(f"=== Workbook: {workbook_name} | Sheet: {sheet_name} ===\n{output}\n{'='*80}\n\n")

    print(f"\n Results for '{language}' saved to '{output_llm_file}'.")

print("\n All processing completed!")
print(f"Prompts saved incrementally to '{OUTPUT_PROMPTS_FILE}'")
print(f"LLM outputs saved in '{OUTPUT_LLM_BASE_DIR}'")
